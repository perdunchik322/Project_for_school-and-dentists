from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidgetItem
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QHeaderView
from openpyxl import load_workbook
from PyQt5 import uic
import sys

# Класс обработчик эксель файлов
class ExcelProcessor:
    def __init__(self, path):  # создание переменныз используемых в дальнейшем
        self.path_to_file = path
        self.file = load_workbook(f"{path}", data_only=True)
        self.ws = self.file["Начисления"]
        self.table_name = "Итог"
        self.total_sum = 0
        self.global_minus = 0
        self.data = dict()
        self.data_for_total = list()

    def get_data(self):  # функция получающая нужные данные из таблицы и обрабатываающая их по тз
        needed_columns = ["G", "H", "I", "J", "Y"]
        for row in range(2, self.ws.max_row + 1):
            row_data = []
            for col in needed_columns:
                cell_value = self.ws[f"{col}{row}"].value
                row_data.append(cell_value)
            if None == row_data[0]:  # отсев глобального минуса(операции в чистый минус)
                self.global_minus += row_data[-1]
            else:
                art = str(row_data[0])
                if art not in self.data:  # создание хеш таблицы для каждого уникального артикула
                    self.data[art] = [0, 0, 0, 0]
                    self.data[art][0] = row_data[1]
                quantity = row_data[2]  # переменные с значениями ячеек, более понятные названия
                sale = row_data[3]
                total = row_data[4]
                self.data[art][3] += total  # s1 из тз
                if sale > 0:  # расчёт кол-ва продаж для артикула
                    self.data[art][1] += quantity
                elif sale < 0:
                    self.data[art][1] -= quantity
        for art in self.data:  # фильтрация тех продаж что ушли в минус
            quantity_art = self.data[art][1]
            if quantity_art <= 0:
                self.global_minus += self.data[art][3]
                self.data[art][2] = -1
            else:
                self.data_for_total.append([art] + self.data[art])
        self.total_sum = sum(row[-1] for row in self.data_for_total)
        for row_ind in range(len(self.data_for_total)):  # распределения минуса на всех
            s1 = self.data_for_total[row_ind][-1]
            self.data_for_total[row_ind][-1] = s1 + s1 / self.total_sum * self.global_minus
            self.data_for_total[row_ind][-2] = self.data_for_total[row_ind][-1] / self.data_for_total[row_ind][-3]
        return self.data_for_total

    def write_data(self):  # создание нового листа в изначальном файле и запись туда обработанной информации
        if self.table_name in self.file.sheetnames:  # проверка на существование записий и чистка лишних(чтобы не плодить листы)
            std = self.file[self.table_name]
            self.file.remove(std)
        ws = self.file.create_sheet(title=self.table_name)  # новый лист с именем "Итог"
        ws.append(["Артикул", "Название", "Кол-во", "Цена, шт", "Итого, So"])  # добавление строки с заголовками

        # Записываем данные
        for row_ind in range(len(self.data_for_total)):
            row = self.data_for_total[row_ind]
            ws.append(row)
        self.file.save(self.path_to_file)


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('ui.ui', self)
        self.path_to_file = ''
        self.center_window()
        self.setup()

    def setup(self):  # подключения тригеров всех кнопок и добавление шорткатов
        self.choose_file_action.setShortcut("Ctrl+C")
        self.save_action.setShortcut("Ctrl+S")
        self.choose_file_action.triggered.connect(self.choose_file)
        self.save_action.triggered.connect(self.write_in_file)

    def choose_file(
            self):  # получение файла и создание экземпляра обработчика файла по полученному из диалог. окна пути
        try:
            fname = QFileDialog.getOpenFileName(self, "Выбрать таблицу", '',
                                                'Таблица (*.xlsx);;Таблица (*.xlsm);;Таблица (*.xls);;Все файлы(*)')
            self.path_to_file = fname[0]
            self.processor = ExcelProcessor(self.path_to_file)
            self.data_for_table = self.processor.get_data()
            self.init_view_table()
        except Exception:
            if fname == ('', ''):
                QMessageBox.information(self, "Справка", "Укажите путь к файлу")
            else:
                QMessageBox.critical(self, "Ошибка", "Ошибка обработки файла")

    def write_in_file(self):  # запись и отображение в статус баре успешность операции
        try:
            self.processor.write_data()
            self.statusBar().showMessage(f'Файл сохранён по пути:{self.path_to_file}')
            self.init_view_table()
        except Exception:
            QMessageBox.critical(self, "Ошибка", "Ошибка обработки файла")

    def init_view_table(self):  # заполнение таблицы внутри окна приложения с результатом программы
        self.view_table.setColumnCount(5)
        self.view_table.setHorizontalHeaderLabels(['Артикул', 'Название', 'Кол-во', 'Цена, шт', 'Итого'])
        self.view_table.setRowCount(len(self.data_for_table))
        header = self.view_table.horizontalHeader()

        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)

        for line_ind, data in enumerate(self.data_for_table):
            self.view_table.setItem(line_ind, 0, QTableWidgetItem(str(data[0])))
            self.view_table.setItem(line_ind, 1, QTableWidgetItem(str(data[1])))
            self.view_table.setItem(line_ind, 2, QTableWidgetItem(str(data[2])))
            self.view_table.setItem(line_ind, 3, QTableWidgetItem(str(data[3])))
            self.view_table.setItem(line_ind, 4, QTableWidgetItem(str(data[4])))

    def center_window(self):
        screen_geometry = self.screen().availableGeometry()  # геометрические данные экрана
        window_geometry = self.frameGeometry()  # геометрические данные окна программы
        center_point = screen_geometry.center()  # получение центра экрана(в пикселях)
        window_geometry.moveCenter(center_point)  # перемещение центра окна в центр экрана в виртуальном прямоугольнике
        self.move(window_geometry.topLeft())  # перемещение реального окна


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
